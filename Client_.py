from config import client_id,accntNmber,password,redirect_uri
from Stream import TDStreamerClient
from urllib.parse import urlparse
from datetime import datetime
from datetime import timedelta
import urllib.parse
import urllib3
import uuid
import os
import json
import requests
import dateutil.parser
import time
import csv
import pandas as pd
import numpy as np
import os
from os import path
from openpyxl import load_workbook




class TDClient():

    def __init__(self,**kwargs):
        self.config = {'consumer_id': client_id,
                       'account_number': accntNmber,
                       'account_password': password,
                       'api_endpoint': 'https://api.tdameritrade.com',
                       'redirect_uri': redirect_uri,
                       'resource': 'https://api.tdameritrade.com',
                       'api_version': '/v1',
                       'cache_state': True,
                       'authenticaiton_url': 'https://auth.tdameritrade.com',
                       'auth_endpoint': 'https://auth.tdameritrade.com' + '/auth?',
                       'token_endpoint': 'https://api.tdameritrade.com' + '/v1' + '/oauth2/token',
                       'refresh_enabled': True
                       }
        self.endpoint_arguments = {'search_instruments': {'projection': ['symbol-search', 'symbol-regex', 'desc-search', 'desc-regex', 'fundamental']},
                                   'get_market_hours': {'markets': ['EQUITY', 'OPTION', 'FUTURE', 'BOND', 'FOREX']},
                                   'get_movers': {'market': ['$DJI', '$COMPX', '$SPX.X'],
                                                  'direction': ['up', 'down'],
                                                  'change': ['value', 'percent']},
                                   'get_user_principals': {'fields': ['streamerSubscriptionKeys', 'streamerConnectionInfo', 'preferences', 'surrogateIds']}
                                  }
        for key in kwargs:
            if key not in self.config:
                print('Warning: the argument {} is an unknown argument'.format(key))
                raise KeyError('Invalid Argument Name.')
        self.config.update(kwargs.items())
        self.state_manager('init')
        self.authstate = False
    def __repr__(self):
        if self.state['loggedin']:
            logged_in_state = 'True'
        else:
            logged_in_state = 'False'
        str_representation = '<TDAmeritrade Client (logged_in = {}, authorized = {})>'.format(logged_in_state, self.authstate)
        return str_representation
    def headers(self, mode=None, token=None) -> dict:
        token = self.state['access_token']
        headers = {'Authorization': 'Bearer {token}'.format(token = self.state['access_token'])}
        if mode == 'application/json':
            headers['Content-Type'] = 'application/json'
        if mode == 'json':
            headers['Content-Type'] = 'application/json'
        return headers
    def api_endpoint(self, endpoint: str, resource: str = None) -> str:
        if resource:
            parts = [resource, self.config['api_version'], endpoint]
        else:
            parts = [self.config['api_endpoint'], self.config['api_version'], endpoint]
        return '/'.join(parts)
    def state_manager(self, action):
        initialized_state = {'access_token': None,
                             'refresh_token': None,
                             'access_token_expires_at': 0,
                             'refresh_token_expires_at': 0,
                             'authorization_url': None,
                             'redirect_code': None,
                             'token_scope': '',
                             'loggedin': False}
        dir_path = os.path.dirname(os.path.realpath(__file__))
        #dir_path = r'C:\Dan\Projects\TD_API\TD-AmeritradeStrategyCode'
        filename = 'TDAmeritradeState.json'
        file_path = os.path.join(dir_path, filename)
        if action == 'init':
            self.state = initialized_state
            if self.config['cache_state'] and os.path.isfile(file_path):
                with open(file_path, 'r') as fileHandle:
                    self.state.update(json.load(fileHandle))
            elif not self.config['cache_state'] and os.path.isfile(os.path.join(dir_path,filename)):
                os.remove(file_path)
        elif action == 'save' and self.config['cache_state']:
            with open(file_path, 'w') as fileHandle:
                json_string = {key:self.state[key] for key in initialized_state}
                json.dump(json_string, fileHandle)
    def login(self):
        if self.config['cache_state']:
            if self.silent_sso():
                self.authstate = 'Authenticated'
                return True
        self.authstate = 'Authenticated'
        payload = {'response_type':'code',
                   'redirect_uri':'http://localhost/',
                   'client_id':self.config['consumer_id'] + '@AMER.OAUTHAP'}
        params = urllib.parse.urlencode(payload)
        url = self.config['auth_endpoint'] + params
        self.state['authorization_url'] = url
        print('Please go to url provided authorize your account: {}'.format(self.state['authorization_url']))
        my_response = input('Paste the full URL resirect here: ')
        self.state['redirect_code'] = my_response
        self.grab_access_token()
    def logout(self):
        self.state_manager('init')
    def grab_access_token(self):
        url_dict = urllib.parse.parse_qs(self.state['redirect_code'])
        url_values = list(url_dict.values())
        url_code = url_values[0][0]
        data = {'grant_type':'authorization_code',
                'client_id':self.config['consumer_id'],
                'access_type':'offline',
                'code':url_code,
                'redirect_uri':'http://localhost/'
               }
        print(url_code)
        token_response = requests.post(url = self.config['token_endpoint'], data=data, verify=True)
        print(token_response)
        self.token_save(token_response)
        if token_response and token_response.ok:
            self.state_manager('save')
    def silent_sso(self):
        if self.token_seconds(token_type='access_token') > 0:
            return True
        elif self.token_seconds(token_type='refresh_token') <= 0:
            return False
        elif self.state['refresh_token'] and self.token_refresh():
            return True
        return False
    def token_refresh(self):
        data = {'grant_type':'refresh_token',
                'client_id':self.config['consumer_id'] + '@AMER.OAUTHAP',
                'refresh_token':self.state['refresh_token'],
                'access_type':'offline'
               }
        response = requests.post(url = self.config['token_endpoint'], data=data, verify=True) 
        if response.status_code == 401:
            print('The Credentials you passed through are invalid.')
            return False
        elif response.status_code == 400:
            print('Validation was unsuccessful.')
            return False
        elif response.status_code == 500:
            print('The TD Server is experiencing an error, please try again later.')
            return False
        elif response.status_code == 403:
            print("You don't have access to this resource, cannot authenticate.")
            return False
        elif response.status_code == 503:
            print("The TD Server can't respond, please try again later.")
            return False
        else:
            self.token_save(response)
            self.state_manager('save')
            return True
    def token_save(self, response):
        json_data = response.json()
        if 'access_token' not in json_data:
            self.logout()
            return False
        self.state['access_token'] = json_data['access_token']
        self.state['refresh_token'] = json_data['refresh_token']
        self.state['loggedin'] = True
        self.state['access_token_expires_at'] = time.time() + int(json_data['expires_in'])
        self.state['refresh_token_expires_at'] = time.time() + int(json_data['refresh_token_expires_in'])
        return True
    def token_seconds(self, token_type = 'access_token'):
        if token_type == 'access_token':
            if not self.state['access_token'] or time.time() >= self.state['access_token_expires_at']:
                return 0
            token_exp = int(self.state['access_token_expires_at'] - time.time())
        elif token_type == 'refresh_token':
            if not self.state['refresh_token'] or time.time() >= self.state['refresh_token_expires_at']:
                return 0
            token_exp = int(self.state['refresh_token_expires_at'] - time.time())
        return token_exp
    def token_validation(self, nseconds = 5):
        if self.token_seconds() < nseconds and self.config['refresh_enabled']:
            self.token_refresh()
    def _create_token_timestamp(self, token_timestamp = None):
        token_timestamp = datetime.strptime(token_timestamp, "%Y-%m-%dT%H:%M:%S%z")
        token_timestamp = int(token_timestamp.timestamp()) * 1000
        return token_timestamp
# CREATE STREAMING SESSION
    def validate_arguments(self, endpoint=None, parameter_name=None, parameter_argument=None):
        parameters_dictionary = self.endpoint_arguments[endpoint]
        parameter_possible_arguments = parameters_dictionary[parameter_name]
        if type(parameter_argument) is list:
            validation_result = [
                argument not in parameter_possible_arguments for argument in parameter_argument]
            if any(validation_result):
                print('\nThe value you passed through is not valid, please choose one of the following valid values: {} \n'.format(
                    ' ,'.join(parameter_possible_arguments)))
                raise ValueError('Invalid Value.')
            elif not any(validation_result):
                return True
        elif parameter_argument not in parameter_possible_arguments:
            print('\nThe value you passed through is not valid, please choose one of the following valid values: {} \n'.upper(
            ).format(' ,'.join(parameter_possible_arguments)))
            raise ValueError('Invalid Value.')
        elif parameter_argument in parameter_possible_arguments:
            return True
    def prepare_arguments_list(self, parameter_list=None):
        if type(parameter_list) is list:
            delimeter = ','
            parameter_list = delimeter.join(parameter_list)
        return parameter_list
    def get_quotes(self, instruments=None):
        self.token_validation()
        merged_headers = self.headers()
        instruments = self.prepare_arguments_list(parameter_list=instruments)
        data = {'apikey': self.config['consumer_id'],
                'symbol': instruments}
        endpoint = '/marketdata/quotes'
        url = self.api_endpoint(endpoint)
        return requests.get(url=url, headers=merged_headers, params=data, verify=True).json()
    def get_user_principals(self, fields=None):
        self.token_validation()
        self.validate_arguments(endpoint='get_user_principals',parameter_name='fields', parameter_argument=fields)
        merged_headers = self.headers()
        fields = self.prepare_arguments_list(parameter_list=fields)
        endpoint = '/userprincipals'
        data = {'fields': fields}
        url = self.api_endpoint(endpoint)
        return requests.get(url=url, headers=merged_headers, params=data, verify=True).json()
    def create_streaming_session(self):
        userPrincipalsResponse = self.get_user_principals(fields = ['streamerConnectionInfo'])
        tokenTimeStamp = userPrincipalsResponse['streamerInfo']['tokenTimestamp']
        socket_url = userPrincipalsResponse['streamerInfo']['streamerSocketUrl']
        tokenTimeStampAsMs = self._create_token_timestamp(token_timestamp=tokenTimeStamp)
        print(tokenTimeStamp, socket_url,tokenTimeStampAsMs)
        credentials = {'userid':userPrincipalsResponse['accounts'][0]['accountId'],
                       'token':userPrincipalsResponse['streamerInfo']['token'],
                       'company':userPrincipalsResponse['accounts'][0]['company'],
                       'segment':userPrincipalsResponse['accounts'][0]['segment'],
                       'cddomain':userPrincipalsResponse['accounts'][0]['accountCdDomainId'],
                       'usergroup':userPrincipalsResponse['streamerInfo']['userGroup'],
                       'accesslevel':userPrincipalsResponse['streamerInfo']['accessLevel'],
                       'authorized':'Y',
                       'timestamp':int(tokenTimeStampAsMs),
                       'appid':userPrincipalsResponse['streamerInfo']['appId'],
                       'acl':userPrincipalsResponse['streamerInfo']['acl'],
                     }
        streaming_session = TDStreamerClient(websocket_url=socket_url, user_principal_data=userPrincipalsResponse,credentials=credentials)
        return streaming_session
#WATCHLIST
#Reads the tickers you have added to the watchlist.csv togather OHLC and streaming price data
    def multiple_symbol_watchlist(self, symbols=None):
        os.chdir('C:\SourceCode\TD-AmeritradeAPI')
        with open('WatchList.csv', newline='') as watchlist:
            WatchList = csv.reader(watchlist, delimiter=',')
            for Symbol in WatchList:
                symbols = Symbol
                return symbols
    def epoch_datetime(self):
        TimeDay = time.strftime('%Y-%m-%d', time.localtime()) 
        TimeSec = time.strftime('%I:%M:%S', time.localtime()) 
        Minus20Day = (datetime.now() - timedelta(days=20))
        #for days in range(1,20,1):
        #    epoch = ((datetime.now() - timedelta(days)).timestamp()) * 1000
        return TimeDay
#INDICATORS
#Interfaces with TD Historical Endpoint to gather OHLC Data called from TDAmaeritrade startegy
    #Have OHLC file to plot Candlestick plots for each symbol as well as daily price changes
    def Historical_Endpoint(self, 
                            symbol:str, 
                            period_type:str=None, 
                            period:str=None,
                            start_date:str=None, 
                            end_date:str=None,
                            frequency_type:str=None,
                            frequency:str=None,
                            extended_hours:bool=True):
        #Historical Data
        # daily proces endpoint
        historicalEndpoint = r'https://api.tdameritrade.com/v1/marketdata/{}/pricehistory'.format(symbol)
        # define a payload
        merged_headers = self.headers()
        historicalPayload = {'apikey':client_id,
                             'period': period,
                             'periodType': period_type,
                             'startDate': start_date,
                             'endDate': end_date,
                             'frequency': frequency,
                             'frequencyType': frequency_type,
                             'needExtendedHoursData': extended_hours
                             }
        # make a request
        historicalContent = requests.get(url=historicalEndpoint, headers=merged_headers, params=historicalPayload)
        print(historicalContent)
        # convert it to a dictionary
        historicalData = historicalContent.json()
        print(historicalData)
        #Store each parameter as a variable and create an array
        Symbol = historicalData['symbol']
        Open = historicalData['candles'][0]['open']
        High = historicalData['candles'][0]['high']
        Low = historicalData['candles'][0]['low']
        Close = historicalData['candles'][0]['close']
        Volume = historicalData['candles'][0]['volume']
        DateTime = historicalData['candles'][0]['datetime'] / 1000
        Day_time = datetime.fromtimestamp(DateTime).strftime('%Y-%m-%d')
        OHLC = [Symbol, Day_time, Open, High, Low, Close, Volume]
        self._write_OHLC_to_csv(OHLC=OHLC, Symbol=Symbol)
#Writes historical data to OHLC file for multiple Symbols called from TDAmeritrade Strategy
    #Create seperate file for each symbol and right over data rather than append (Done)
    #Data files are sreated in the Data Folder(Done)
        #Create a different folder for each day(Done)
    #CSV Dates are incorrect (Accounts for Weekends)
    def _write_OHLC_to_csv(self, OHLC, Symbol):
        Date = self.epoch_datetime()
        if path.exists('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC'):
            os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
            with open((Symbol + '_' + 'OHLC' + '_' + Date + '.csv'), mode='a+', newline='') as OHLC_file:           
                OHLC_writer = csv.writer(OHLC_file)
                historicalData = OHLC
                if os.path.getsize((Symbol + '_' + 'OHLC' + '_' + Date + '.csv')) == 0:
                    OHLC_writer.writerow(['Symbol','Date','Open','High','Low','Close','Volume'])
                OHLC_writer.writerow(historicalData)
                os.chdir('C:\SourceCode\TD-AmeritradeAPI')
        else:
            os.mkdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date)
            os.mkdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
            os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
            with open((Symbol + '_' + 'OHLC' + '_' + Date + '.csv'), mode='a+', newline='') as OHLC_file:           
                OHLC_writer = csv.writer(OHLC_file)
                historicalData = OHLC
                if os.path.getsize((Symbol + '_' + 'OHLC' + '_' + Date + '.csv')) == 0:
                    OHLC_writer.writerow(['Symbol','Date','Open','High','Low','Close','Volume'])
                OHLC_writer.writerow(historicalData)
                os.chdir('C:\SourceCode\TD-AmeritradeAPI')
#INDICATORS
    def _SMA_(self, symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime()) 
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        SMA = {}
        for Ticker in symbol:
            SMA[Ticker] = pd.read_csv(Ticker + '_' + 'OHLC' + '_' + Date + '.csv')
            SMA[Ticker] = SMA[Ticker].iloc[:,5].rolling(window=1).mean()
            SMA[Ticker + ' ' + 'SMA'] = SMA.pop(Ticker) 
            dfSMA_data = pd.DataFrame(SMA)
        return dfSMA_data
    def _SMA_toCSV(self, symbol=None, SimpleMovingAverage=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        for Ticker in symbol:
            df = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            df = df.merge(SimpleMovingAverage[[Ticker + ' ' + 'SMA']], left_index=True, right_index=True)
            df.to_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'), index=False)
    def fiftyDaySMA(self,symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        fiftyDaySMA = {}
        for ticker in symbol:
            fiftyDaySMA[ticker] = pd.read_csv((ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            fiftyDaySMA[ticker] = fiftyDaySMA[ticker].iloc[:16,7].rolling(window=28, min_periods=0).mean()
            #fiftyDaySMA[ticker + ' ' + '50DaySMA'] = fiftyDaySMA.pop(ticker)
        return fiftyDaySMA
    def twentyDaySMA(self,symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        twentyDaySMA = {}
        for ticker in symbol:
            twentyDaySMA[ticker] = pd.read_csv((ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            twentyDaySMA[ticker] = twentyDaySMA[ticker].iloc[:11,7].rolling(window=10, min_periods=0).mean()
            #twentyDaySMA[ticker + ' ' + '20DaySMA'] = twentyDaySMA.pop(ticker)
        return twentyDaySMA
    def spanTwelveEMA(self,symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        spanTwelveEMA = {}
        for ticker in symbol:
            spanTwelveEMA[ticker] = pd.read_csv((ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            spanTwelveEMA[ticker] = (spanTwelveEMA[ticker].iloc[::-1,5]).ewm(span=12,adjust=False).mean()
            spanTwelveEMA[ticker + ' ' + 'spanTwelveEMA'] = spanTwelveEMA.pop(ticker) 
            df_spanTwelveEMA = pd.DataFrame(spanTwelveEMA)
        return df_spanTwelveEMA
    def spanTwntySixEMA(self,symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        spanTwntySixEMA = {}
        for ticker in symbol:
            spanTwntySixEMA[ticker] = pd.read_csv((ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            spanTwntySixEMA[ticker] = (spanTwntySixEMA[ticker].iloc[::-1,5]).ewm(span=26, adjust=False).mean()
            spanTwntySixEMA[ticker + ' ' + 'spanTwntySixEMA'] = spanTwntySixEMA.pop(ticker) 
            df_spanTwntySixEMA = pd.DataFrame(spanTwntySixEMA)
        return df_spanTwntySixEMA
    def _MACD_(self,symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        spanTwelveEMA = self.spanTwelveEMA(symbol=symbol)
        spanTwntySixEMA = self.spanTwntySixEMA(symbol=symbol)
        MACD = {}
        for ticker in symbol:
            MACD[ticker] = spanTwelveEMA[ticker + ' ' + 'spanTwelveEMA'] - spanTwntySixEMA[ticker + ' ' + 'spanTwntySixEMA']
            MACD[ticker] = MACD.pop(ticker) 
            df_MACD = pd.DataFrame(MACD)
        return df_MACD
    def _MACD_Tickers(self,symbol):
        MACD_Tickers = {}
        dfMACD = self._MACD_(symbol=symbol)
        MACD_Tickers = dfMACD.to_dict()
        return MACD_Tickers
    def _MACD_signalTickers(self, symbol):
        MACD_signalTickers = {}
        dfMACD_Signal = self.MACD_Signal(symbol=symbol)
        MACD_signalTickers = dfMACD_Signal.to_dict()
        return MACD_signalTickers
    def MACD_Signal(self,symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        MACD_Signal = {}
        for ticker in symbol:
            MACD_Signal[ticker] = pd.read_csv((ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            MACD_Signal[ticker] = (MACD_Signal[ticker].iloc[::-1,8])-(MACD_Signal[ticker].iloc[::-1,9])
            MACD_Signal[ticker] = (MACD_Signal[ticker]).ewm(span=9, adjust=False).mean()
            MACD_Signal[ticker] = MACD_Signal.pop(ticker) 
            df_MACDsignal = pd.DataFrame(MACD_Signal)
        return df_MACDsignal
    def _EMA_toCSV(self, symbol, spantwelveEMA=None, spanTwntySixEMA=None, _MACD_=None, MACD_Signal=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        for Ticker in symbol:
            df = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            df = df.merge(spantwelveEMA[[Ticker + ' ' + 'spanTwelveEMA']], left_index=True, right_index=True)
            df = df.merge(spanTwntySixEMA[[Ticker + ' ' + 'spanTwntySixEMA']], left_index=True, right_index=True)
            df = df.merge(_MACD_[[Ticker]], left_index=True, right_index=True)
            df.to_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'), index=False)
    def _MACD_SignaltoCSV(self, symbol=None, MACD_Signal=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        for Ticker in symbol:
            df = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            df = df.merge(MACD_Signal[[Ticker]], left_index=True, right_index=True)
            df.to_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'), index=False)
    def Momentum(self, symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        prevClose = {}
        lookBack = {}
        momentum = {}
        for Ticker in symbol:
            prevClose[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            prevClose[Ticker] = prevClose[Ticker].iloc[0,5]
            lookBack[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            lookBack[Ticker] = lookBack[Ticker].iloc[11,5]
            momentum[Ticker] = prevClose[Ticker] - lookBack[Ticker]
            #momentum[Ticker] = momentum.pop(Ticker)
            df_momentum = pd.DataFrame([momentum])
            df_momentum = df_momentum.transpose()
            df_momentum.rename(columns={0:'Momentum'}, inplace=True)
        return df_momentum
    def stdev(self, symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        stdev = {}
        for Ticker in symbol:
            stdev[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            stdev[Ticker] = stdev[Ticker].iloc[:,5]
            stdev[Ticker] = np.std(stdev[Ticker])
            #stdev[Ticker] = stdev.pop(Ticker)
            df_stdev = pd.DataFrame([stdev])
            df_stdev = df_stdev.transpose()
            #df_stdev.rename(columns={0:'stdev'}, inplace=True)
        return df_stdev
    def meanClose(self, symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        meanClose = {}
        for Ticker in symbol:
            meanClose[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            meanClose[Ticker] = meanClose[Ticker].iloc[:,5].mean()
            df_meanClose = pd.DataFrame([meanClose])
            df_meanClose = df_meanClose.transpose()
            df_meanClose.rename(columns={0:'MeanClose'}, inplace=True)
        return df_meanClose
    def meanHigh(self, symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        meanHigh = {}
        for Ticker in symbol:
            meanHigh[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            meanHigh[Ticker] = meanHigh[Ticker].iloc[:,3].mean()
            df_meanHigh = pd.DataFrame([meanHigh])
            df_meanHigh = df_meanHigh.transpose()
            df_meanHigh.rename(columns={0:'MeanHigh'}, inplace=True)
        return df_meanHigh
    def meanLow(self, symbol):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        meanLow = {}
        for Ticker in symbol:
            meanLow[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            meanLow[Ticker] = meanLow[Ticker].iloc[:,4].mean()
            df_meanLow = pd.DataFrame([meanLow])
            df_meanLow = df_meanLow.transpose()
            #df_meanLow.rename(columns={0:'MeanLow'}, inplace=True)
        return df_meanLow
    def twoSigmaUp(self, symbol, stdev=None):
        stdev = self.stdev(symbol=symbol)
        meanLow = self.meanLow(symbol=symbol)
        twoStdev = {key: value * 1.5 for key, value in stdev.items()}
        df_twoStdev = pd.DataFrame(twoStdev)
        twoSigmaUp = df_twoStdev.radd(meanLow)
        twoSigmaUp.rename(columns={0:'twoSigmaUp'}, inplace=True)
        return twoSigmaUp
    def twoSigmaDown(self, symbol, stdev=None):
        stdev = self.stdev(symbol=symbol)
        meanClose = self.meanClose(symbol=symbol)
        twoStdev = {key: value * 2 for key, value in stdev.items()}
        df_twoStdev = pd.DataFrame(twoStdev)
        twoSigmaDown = df_twoStdev.rsub(meanClose)
        twoSigmaDown.rename(columns={0:'twoSigmaDown'}, inplace=True)
        return twoSigmaDown
    def two_stdevSell(self, symbol=None):
        twoSigmaUp = self.twoSigmaUp(symbol=symbol)
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        Low = {}
        for Ticker in symbol:
            Low[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            Low[Ticker] = Low[Ticker].iloc[0,4]
            df_Low = pd.DataFrame([Low])
            df_Low = df_Low.transpose()
            df_Low.rename(columns={0:'MeanLow'}, inplace=True)
        df_twoSigmaUpSell = df_Low.merge(twoSigmaUp, left_index=True, right_index=True)
        return df_twoSigmaUpSell
    def CloseMeanSell(self, symbol=None):
        meanClose = self.meanClose(symbol=symbol)
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        close = {}
        for Ticker in symbol:
            close[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            close[Ticker] = close[Ticker].iloc[0,5]
            df_close = pd.DataFrame([close])
            df_close = df_close.transpose()
            df_close.rename(columns={0:'Close'}, inplace=True)
        df_CloseMean = df_close.merge(meanClose, left_index=True, right_index=True)
        return df_CloseMean
    def CloseTrend(self, symbol=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'OHLC')
        close = {}
        closeLookback = {}
        for Ticker in symbol:
            close[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            close[Ticker] = close[Ticker].iloc[0,5]
            df_close = pd.DataFrame([close])
            df_close = df_close.transpose()
            df_close.rename(columns={0:'Close'}, inplace=True)
            closeLookback[Ticker] = pd.read_csv((Ticker + '_' + 'OHLC' + '_' + Date + '.csv'))
            closeLookback[Ticker] = closeLookback[Ticker].iloc[4,5]
            df_closeLookback = pd.DataFrame([closeLookback])
            df_closeLookback = df_closeLookback.transpose()
            df_closeLookback.rename(columns={0:'CloseLookback'}, inplace=True)
        df_closeTrend = df_close.merge(df_closeLookback, left_index=True, right_index=True)
        return df_closeTrend
#BUY/SELL SIGNALS
    def SMABuyTickers(self,symbol=None):
        fiftyDaySMA = self.fiftyDaySMA(symbol=symbol)
        fiftyDaySMA_Values = pd.DataFrame.from_dict(fiftyDaySMA, orient='index')
        fiftyDaySMA_Values = pd.DataFrame(fiftyDaySMA_Values)
        fiftyDaySMA_Values = fiftyDaySMA_Values.iloc[:,-1]
        fiftyDaySMA_Values = fiftyDaySMA_Values.to_frame()
        twentyDaySMA = self.twentyDaySMA(symbol=symbol)
        twentyDaySMA_Values = pd.DataFrame.from_dict(twentyDaySMA, orient='index')
        twentyDaySMA_Values = pd.DataFrame(twentyDaySMA_Values)
        twentyDaySMA_Values = twentyDaySMA_Values.iloc[:,-1]
        twentyDaySMA_Values = twentyDaySMA_Values.to_frame()
        df_SMA = twentyDaySMA_Values.merge(fiftyDaySMA_Values, left_index=True, right_index=True)
        df_SMA.rename(columns={10:'fastSMA',15:'slowSMA'}, inplace=True)
        BuyTickers = df_SMA[df_SMA['fastSMA'] > df_SMA['slowSMA']].index
        BuyTickers = BuyTickers.tolist()
        return BuyTickers
    def MACD_buyTickers (self, symbol=None):
        MACD_Tickers = self._MACD_Tickers(symbol=symbol)
        MACD_Values = pd.DataFrame.from_dict(MACD_Tickers, orient='index')
        prevMACD = MACD_Values[0]
        dfprevMACD = pd.DataFrame(prevMACD)
        MACD_signalTickers = self._MACD_signalTickers(symbol=symbol)
        MACD_Signal_Values = pd.DataFrame.from_dict(MACD_signalTickers, orient='index')
        prevMACD_Signal = MACD_Signal_Values[0]
        dfprevMACD_Signal = pd.DataFrame(prevMACD_Signal)
        df_MACD = dfprevMACD.merge(dfprevMACD_Signal, left_index=True, right_index=True)
        df_MACD.rename(columns={'0_x':'MACD','0_y':'MACD Signal'}, inplace=True)
        print(df_MACD)
        buyTickers = df_MACD[df_MACD['MACD'] > df_MACD['MACD Signal']].index
        buyTickers = buyTickers.tolist()
        return buyTickers
    def MomentumBuyTickers(self, symbol=None):
        Momentum = self.Momentum(symbol=symbol)
        MomentumBuy = Momentum[Momentum['Momentum'] > 0].index
        MomentumBuy = MomentumBuy.tolist()
        return MomentumBuy
    def SMA_SellTickers(self,symbol=None):
        fiftyDaySMA = self.fiftyDaySMA(symbol=symbol)
        fiftyDaySMA_Values = pd.DataFrame.from_dict(fiftyDaySMA, orient='index')
        fiftyDaySMA_Values = pd.DataFrame(fiftyDaySMA_Values)
        fiftyDaySMA_Values = fiftyDaySMA_Values.iloc[:,-1]
        fiftyDaySMA_Values = fiftyDaySMA_Values.to_frame()
        twentyDaySMA = self.twentyDaySMA(symbol=symbol)
        twentyDaySMA_Values = pd.DataFrame.from_dict(twentyDaySMA, orient='index')
        twentyDaySMA_Values = pd.DataFrame(twentyDaySMA_Values)
        twentyDaySMA_Values = twentyDaySMA_Values.iloc[:,-1]
        twentyDaySMA_Values = twentyDaySMA_Values.to_frame()
        df_SMA = twentyDaySMA_Values.merge(fiftyDaySMA_Values, left_index=True, right_index=True)
        df_SMA.rename(columns={10:'fastSMA',15:'slowSMA'}, inplace=True)
        print(df_SMA)
        SellTickers = df_SMA[df_SMA['fastSMA'] < df_SMA['slowSMA']].index
        SellTickers = SellTickers.tolist()
        return SellTickers
    def MACD_SellTickers(self,symbol=None):
        MACD_Tickers = self._MACD_Tickers(symbol=symbol)
        MACD_Values = pd.DataFrame.from_dict(MACD_Tickers, orient='index')
        prevMACD = MACD_Values[0]
        dfprevMACD = pd.DataFrame(prevMACD)
        MACD_signalTickers = self._MACD_signalTickers(symbol=symbol)
        MACD_Signal_Values = pd.DataFrame.from_dict(MACD_signalTickers, orient='index')
        prevMACD_Signal = MACD_Signal_Values[0]
        dfprevMACD_Signal = pd.DataFrame(prevMACD_Signal)
        df_MACD = dfprevMACD.merge(dfprevMACD_Signal, left_index=True, right_index=True)
        df_MACD.rename(columns={'0_x':'MACD','0_y':'MACD Signal'}, inplace=True)
        SellTickers = df_MACD[df_MACD['MACD'] < df_MACD['MACD Signal']].index
        SellTickers = SellTickers.tolist()
        return SellTickers
    def readStream(self, position=None):
        Date = time.strftime('%Y-%m-%d', time.localtime())
        os.chdir('C:\SourceCode\TD-AmeritradeAPI\Data' + '\\' + Date + '\\' + 'StreamData')
        streamData = {}
        streamQuote = {}
        askPrice = {}
        for Ticker in position:
            while True:
                if path.exists(Ticker + '_' + 'Stream' + '_' + Date + '.csv'):
                    os.path.isfile((Ticker + '_' + 'Stream' + '_' + Date + '.csv'))
                    streamData[Ticker] = pd.read_csv((Ticker + '_' + 'Stream' + '_' + Date + '.csv'),names=['Symbol','AskPrice','Time'])
                    streamQuote[Ticker] = streamData[Ticker].iloc[-1] 
                    askPrice[Ticker] = streamQuote[Ticker]['AskPrice']
                    break
                    #return askPrice
                else:
                    print('Waiting for quote price.')
                    time.sleep(30)
            #askPrice = pd.DataFrame(data=askPrice, index=askPrice.keys())
        return askPrice
    def two_stdevSellTickers(self, symbol=None):
        two_stdevSell = self.two_stdevSell(symbol=symbol)
        two_stdevSellTickers = two_stdevSell[two_stdevSell['MeanLow'] > two_stdevSell['twoSigmaUp']].index
        two_stdevSellTickers = two_stdevSellTickers.tolist()
        return two_stdevSellTickers
    def CloseMeanSellTickers(self, symbol=None):
        CloseMeanSell = self.CloseMeanSell(symbol=symbol)
        CloseMeanSellTickers = CloseMeanSell[CloseMeanSell['Close'] < CloseMeanSell['MeanClose']].index
        CloseMeanSellTickers = CloseMeanSellTickers.tolist()
        return CloseMeanSellTickers
    def CloseMeanTrendSellTickers(self, symbol=None):
        CloseMeanTrend = self.CloseTrend(symbol=symbol)
        CloseMeanTrendSellTickers = CloseMeanTrend[CloseMeanTrend['Close'] < CloseMeanTrend['CloseLookback']].index
        CloseMeanTrendSellTickers = CloseMeanTrendSellTickers.tolist()
        return CloseMeanTrendSellTickers
#Account Info
    def accounts(self, accntNmber=None):
        AccntPayload = {'fields':'positions',
                        'apikey':client_id
                       }
        merged_headers = self.headers()
        getAccntsEndpoint = r'https://api.tdameritrade.com/v1/accounts/{}'.format(accntNmber)
        AccntContent = requests.get(url=getAccntsEndpoint, headers=merged_headers, params=AccntPayload)
        AccntPositions = AccntContent.json()
        return AccntPositions
    def BuyingPower(self,accntNmber=None):
        accntInfo = self.accounts(accntNmber)
        buyingPower = accntInfo['securitiesAccount']['projectedBalances']['availableFunds']
        buyingPower = float(buyingPower)
        return buyingPower
    def accntAssets(self,accntNmber=None,symbol=None):
        accntInfo = self.accounts(accntNmber=accntNmber)
        assets = {}
        lenAssets = len(symbol)
        try:
            for i in range(lenAssets):
                assets[i] = accntInfo['securitiesAccount']['positions'][i]['instrument']['symbol']  
                data = assets.values()
                Positions = list(data)
        except IndexError:
            'N/A'
        return Positions
    def assetQuantity(self,accntNmber=None,symbol=None):
        accntInfo = self.accounts(accntNmber=accntNmber)
        quantity = {}
        lenAssets = len(symbol)
        try:
            for i in range(lenAssets):
                quantity[i] = accntInfo['securitiesAccount']['positions'][i]['longQuantity']  
                data = quantity.values()
                Quantity = list(data)
        except IndexError:
            'N/A'
        return Quantity
    def TDA_Portfolio(self, symbol, accntNmber):
        Assets = self.accntAssets(accntNmber=accntNmber, symbol=symbol)
        Positions = self.assetQuantity(accntNmber=accntNmber, symbol=symbol)
        dfAssets = pd.DataFrame(Assets)
        dfPositions = pd.DataFrame(Positions)
        dfPortfolio = dfAssets.merge(dfPositions, left_index=True, right_index=True)
        dfPortfolio.rename(columns={'0_x':'Ticker','0_y':'Quantity'}, inplace=True)
        return dfPortfolio
    def TDA_Watchlists(self, accntNmber=None):
        WatchlistPayload = {'apikey':client_id}
        merged_headers = self.headers()
        getWatchlistEndpoint = r'https://api.tdameritrade.com/v1/accounts/{}/watchlists/{}'.format(accntNmber,1466022388)
        WatchlistContent = requests.get(url=getWatchlistEndpoint, headers=merged_headers, params=WatchlistPayload)
        Watchlist = WatchlistContent.json()
        return Watchlist
    def watchlistSymbols(self, accntNmber=None):
        watchlist = self.TDA_Watchlists(accntNmber=accntNmber)
        watchlistSymbols = {}
        try:
            #Dont like hard coded values
            for i in range(100):
                watchlistSymbols[i] = watchlist['watchlistItems'][i]['instrument']['symbol']
                data = watchlistSymbols.values()
                Symbols = list(data)
        except IndexError:
            'N/A'
        return Symbols
    def Portfolio_toExcelRTD(self, symbol=None, accntNmber=None):
        os.chdir('C:\SourceCode\TD-AmeritradeAPI')
        portfolio = self.accntAssets(symbol=symbol,accntNmber=accntNmber)
        excelPortfolioReader = pd.read_excel(r'Portfolio.xlsm', sheet_name='EquityPositions')
        positions = pd.DataFrame(excelPortfolioReader)
        positions = positions['Stock']
        stockPositions = list(positions)
        stockAppend = [i for i in portfolio if i not in stockPositions]
        stockAppend = pd.DataFrame(stockAppend)
        book = load_workbook('Portfolio.xlsm', keep_vba=True)
        writer = pd.ExcelWriter('Portfolio.xlsm', engine='openpyxl', mode='a')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        stockAppend.to_excel(writer, startrow=len(stockPositions)+1, sheet_name='EquityPositions', header=False, index=False)
        writer.save()
    def Watchlist_toExcelRTD(self, symbol=None, accntNmber=None):
        os.chdir('C:\SourceCode\TD-AmeritradeAPI')
        watchlistEquity = self.watchlistSymbols(accntNmber=accntNmber)
        readerExcel = pd.read_excel(r'Portfolio.xlsm')
        watchlistPositions = pd.DataFrame(readerExcel)
        watchlistPositions = watchlistPositions['Stock']
        watchlistStocks = list(watchlistPositions)
        tickerAppend = [i for i in watchlistEquity if i not in watchlistStocks]
        tickerAppend = pd.DataFrame(tickerAppend)
        book = load_workbook('Portfolio.xlsm', keep_vba=True)
        writer = pd.ExcelWriter('Portfolio.xlsm', engine='openpyxl', mode='a')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        tickerAppend.to_excel(writer, startrow=len(watchlistStocks)+1, sheet_name='EquityWatchlist', header=False, index=False)
        writer.save()
#Orders
    def shareNum_buy(self, positions=None, position=None):
        quotePrice = self.readStream(position=position)      
        quotePrice = [float(value) for value in quotePrice.values()]
        if quotePrice <= [1.00]:
            shares = 300
        elif quotePrice <= [2.00]:
            shares = 200
        elif quotePrice <= [3.00]:
            shares = 100
        elif quotePrice <= [4.00]:
            shares = 50
        elif quotePrice <= [5.00]:
            shares = 20
        elif quotePrice <= [6.00]:
            shares = 10
        else:
            shares = 0
        return shares
    def BuyMarketOrder(self, shares:str, ticker:str):
        Order = {'orderType': 'MARKET',
                 'session': 'NORMAL',
                 'duration': 'DAY',
                 'orderStrategyType': 'SINGLE',
                 'orderLegCollection': [{'instruction': 'Buy',
                                                        'quantity': shares,
                                                        'instrument': {'symbol': ticker,
                                                                       'assetType': 'EQUITY'
                                                                      }
                                        }
                                       ]
                }
        placeOrder = json.dumps(Order)
        return placeOrder
    def SellMarketOrder(self, shares:str, ticker:str):
        Order = {'orderType': 'MARKET',
                 'session': 'NORMAL',
                 'duration': 'DAY',
                 'orderStrategyType': 'SINGLE',
                 'orderLegCollection': [{'instruction': 'Sell',
                                                        'quantity': shares,
                                                        'instrument': {'symbol': ticker,
                                                                       'assetType': 'EQUITY'
                                                                      }
                                        }
                                       ]
                }
        sellPositions = json.dumps(Order, indent=4)
        return sellPositions
    def place_order(self, accntNmber=None, mode=None, shares=None, ticker=None):
        headers = self.headers(mode='json')
        orderData = self.BuyMarketOrder(shares=shares, ticker=ticker)
        print(orderData)
        orderEndpoint = r'https://api.tdameritrade.com/v1/accounts/{}/orders'.format(accntNmber)
        PlaceOrder = requests.post(url=orderEndpoint, headers=headers, data=orderData)
        return PlaceOrder
    def sellPositions(self, accntNmber=None, mode=None, shares=None, ticker=None):
        headers = self.headers(mode='json')
        orderData = self.SellMarketOrder(shares=shares, ticker=ticker)
        print(orderData)
        orderEndpoint = r'https://api.tdameritrade.com/v1/accounts/{}/orders'.format(accntNmber)
        SellOrder = requests.post(url=orderEndpoint, headers=headers, data=orderData)
        return SellOrder
    def getOrders(self, accntNmber=None):
        fromEnteredTime = (datetime.now() - timedelta(hours=120)).strftime('%Y-%m-%d')
        toEnteredTime = time.strftime('%Y-%m-%d', time.localtime())
        OrdersPayload = {
                         'apikey':client_id,
                         'fromEnteredTime':fromEnteredTime,
                         'toEnteredTime':toEnteredTime,
                         'maxResults':10,
                         'status':'FILLED'
                       }
        merged_headers = self.headers()
        getOrdersEndpoint = r'https://api.tdameritrade.com/v1/accounts/{}/orders'.format(accntNmber)
        ordersContent = requests.get(url=getOrdersEndpoint, headers=merged_headers, params=OrdersPayload)
        orders = ordersContent.json()
        return orders
    def ordersExcel(self, accntNmber=None):
        orderHistory = self.getOrders(accntNmber=accntNmber)
        quantity = {}
        order = {}
        position = {}
        try:
            #Dont like hard coded values
            for i in range(10):
                quantity[i] = orderHistory[i]['orderLegCollection'][i]['quantity']
                order[i] = orderHistory[i]['orderLegCollection'][i]['instruction']
                position[i] = orderHistory[i]['orderLegCollection'][i]['instrument']['symbol']
                print(order)
                print(position)
        except IndexError:
            'N/A'
        return quantity

   



        


            



        
            
      

